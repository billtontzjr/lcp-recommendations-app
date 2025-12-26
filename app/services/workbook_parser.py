"""Excel workbook parsing service."""
from openpyxl import load_workbook
from datetime import datetime


class WorkbookParseError(Exception):
    """Error parsing workbook."""
    pass


class NoItemsSelectedError(Exception):
    """No items selected in workbook."""
    pass


class MissingPatientInfoError(Exception):
    """Required patient info missing."""
    pass


def parse_workbook(file_path):
    """
    Parse the Master Workbook and extract all data.

    Args:
        file_path: Path to the .xlsm/.xlsx file

    Returns:
        Dictionary with patient_info, items, pfr_lookup, apc_lookup
    """
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise WorkbookParseError(f"Could not open workbook: {str(e)}")

    result = {
        'patient_info': parse_patient_info(wb),
        'items': parse_master_items(wb),
        'pfr_lookup': parse_pfr_sheet(wb),
        'apc_lookup': parse_apc_sheet(wb),
    }

    if not result['items']:
        raise NoItemsSelectedError("No items selected in workbook (Column A 'Check' = True)")

    return result


def parse_patient_info(wb):
    """
    Parse patient information from 'Patient Info' sheet.

    Sheet structure:
        Row 4: Date of Report (Column E)
        Row 5: Client Name (Column E)
        Row 6: Date of Birth (Column E)
        Row 7: Age (Column E)
        Row 8: Date of Injury (Column E)
        Row 9: Life Expectancy (Column E)
        Row 10: Age Initiated (Column E)
        Row 11: Geographic Multiplier (Column E)
        Row 12: City and State (Column E)
        Row 13: Zipcode (Column E)
        Row 14: Source of Referral (Column E)
    """
    if 'Patient Info' not in wb.sheetnames:
        raise MissingPatientInfoError("'Patient Info' sheet not found in workbook")

    ws = wb['Patient Info']

    def get_cell(row, col='E'):
        """Get cell value, handling dates."""
        val = ws[f'{col}{row}'].value
        if isinstance(val, datetime):
            return val
        return val

    patient_info = {
        'date_of_report': get_cell(4),
        'patient_name': get_cell(5),
        'date_of_birth': get_cell(6),
        'age': get_cell(7),
        'date_of_injury': get_cell(8),
        'life_expectancy': get_cell(9),
        'age_initiated': get_cell(10),
        'geographic_multiplier': get_cell(11) or 1.0,
        'city_state': get_cell(12),
        'zipcode': get_cell(13),
        'referring_attorney': get_cell(14),
    }

    # Validate required fields
    if not patient_info['patient_name']:
        raise MissingPatientInfoError("Patient name is required")

    # Calculate until_age if possible
    if patient_info['age_initiated'] and patient_info['life_expectancy']:
        try:
            patient_info['until_age'] = int(patient_info['age_initiated']) + int(float(patient_info['life_expectancy']))
        except (ValueError, TypeError):
            patient_info['until_age'] = None

    return patient_info


def parse_master_items(wb):
    """
    Parse selected items from 'Master' sheet.

    Sheet structure:
        Row 3: Headers
        Column A: Check (Boolean - True if selected)
        Column B: Main Category
        Column C: Item
        Column D: Subcategory
        Column E: Service Description
        Column F: Code Type (PFR, APC, DRG)
        Column G: Code (CPT code or codes)
        Column H: Cost (pre-calculated or to be looked up)
        Column I: Frequency
        Column J: Source
        Column K: Rationale

        Data starts Row 6
    """
    if 'Master' not in wb.sheetnames:
        raise WorkbookParseError("'Master' sheet not found in workbook")

    ws = wb['Master']
    items = []

    # Start from row 6, continue until empty
    row = 6
    while True:
        check_value = ws[f'A{row}'].value

        # Stop if we hit an empty row
        if ws[f'B{row}'].value is None and ws[f'C{row}'].value is None:
            break

        # Only include checked items
        if check_value is True or str(check_value).upper() == 'TRUE':
            item = {
                'category': ws[f'B{row}'].value or '',
                'item': ws[f'C{row}'].value or '',
                'subcategory': ws[f'D{row}'].value or '',
                'service_description': ws[f'E{row}'].value or '',
                'code_type': ws[f'F{row}'].value or '',
                'code': ws[f'G{row}'].value or '',
                'cost': ws[f'H{row}'].value,
                'frequency': ws[f'I{row}'].value or '',
                'source': ws[f'J{row}'].value or '',
                'rationale': ws[f'K{row}'].value or '',
            }
            items.append(item)

        row += 1

        # Safety limit
        if row > 1000:
            break

    return items


def parse_pfr_sheet(wb):
    """
    Parse PFR (Professional Fee) pricing sheet.

    Sheet structure:
        Column A: CPT Code
        Column B: P75 Price
    """
    if 'PFR' not in wb.sheetnames:
        return {}

    ws = wb['PFR']
    lookup = {}

    for row in range(2, ws.max_row + 1):
        cpt = ws[f'A{row}'].value
        price = ws[f'B{row}'].value

        if cpt and price:
            # Handle multiple CPT codes in one cell
            cpt_str = str(cpt).strip()
            try:
                price_val = float(price)
                lookup[cpt_str] = price_val
            except (ValueError, TypeError):
                continue

    return lookup


def parse_apc_sheet(wb):
    """
    Parse APC (Ambulatory Payment Classification) pricing sheet.

    Sheet structure:
        Column A: CPT Code
        Column B: Facility Charge
    """
    if 'APC' not in wb.sheetnames:
        return {}

    ws = wb['APC']
    lookup = {}

    for row in range(2, ws.max_row + 1):
        cpt = ws[f'A{row}'].value
        charge = ws[f'B{row}'].value

        if cpt and charge:
            cpt_str = str(cpt).strip()
            try:
                charge_val = float(charge)
                lookup[cpt_str] = charge_val
            except (ValueError, TypeError):
                continue

    return lookup
